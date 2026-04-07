from datetime import date, datetime, time
from io import BytesIO
from xml.sax.saxutils import escape
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from django.db.models import Q, Sum
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from rest_framework import status
from rest_framework.exceptions import ValidationError
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.inventory.models import Delivery, InternalTransfer, Receipt, StockAdjustment, StockLedger
from apps.products.models import Product
from apps.users.models import User
from apps.users.permissions import IsManagerOrStaff, IsManagerOrSuperUser


NO_DATA_DETAIL = 'No data found for the selected filters.'

TITLE_FONT = Font(name='Calibri', size=17, bold=True, color='1F2937')
META_FONT = Font(name='Calibri', size=10, italic=True, color='475569')
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='1F2937')
BODY_FONT = Font(name='Calibri', size=10, color='111827')

TITLE_ALIGNMENT = Alignment(horizontal='center', vertical='center')
LEFT_ALIGNMENT = Alignment(horizontal='left', vertical='center', wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

HEADER_FILL = PatternFill(fill_type='solid', fgColor='DCE6F1')
ROW_FILL_EVEN = PatternFill(fill_type='solid', fgColor='FFFFFF')
ROW_FILL_ODD = PatternFill(fill_type='solid', fgColor='F8FAFC')

THIN_SIDE = Side(style='thin', color='CBD5E1')
TABLE_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

STATUS_STYLE_MAP = {
	'done': {
		'fill': PatternFill(fill_type='solid', fgColor='D1FAE5'),
		'font': Font(name='Calibri', size=10, bold=True, color='065F46'),
	},
	'waiting': {
		'fill': PatternFill(fill_type='solid', fgColor='FFEDD5'),
		'font': Font(name='Calibri', size=10, bold=True, color='9A3412'),
	},
	'draft': {
		'fill': PatternFill(fill_type='solid', fgColor='E2E8F0'),
		'font': Font(name='Calibri', size=10, bold=True, color='334155'),
	},
	'cancelled': {
		'fill': PatternFill(fill_type='solid', fgColor='FEE2E2'),
		'font': Font(name='Calibri', size=10, bold=True, color='991B1B'),
	},
}


def _query_param(request, *keys):
	for key in keys:
		value = request.query_params.get(key)
		if value not in (None, ''):
			return value
	return None


def _parse_datetime_filter(raw_value, *, end_of_day=False):
	parsed_datetime = parse_datetime(raw_value)
	if parsed_datetime is not None:
		if timezone.is_naive(parsed_datetime):
			parsed_datetime = timezone.make_aware(
				parsed_datetime,
				timezone.get_current_timezone(),
			)
		return parsed_datetime

	parsed_date = parse_date(raw_value)
	if parsed_date is not None:
		boundary = time.max if end_of_day else time.min
		combined = datetime.combine(parsed_date, boundary)
		return timezone.make_aware(combined, timezone.get_current_timezone())

	raise ValidationError(
		{'detail': f'Invalid date filter "{raw_value}". Use YYYY-MM-DD or ISO datetime.'}
	)


def _apply_date_range_filter(queryset, request, field_name='created_at'):
	start_value = _query_param(request, 'date_from', 'start_date', 'from_date')
	end_value = _query_param(request, 'date_to', 'end_date', 'to_date')

	start_datetime = (
		_parse_datetime_filter(start_value, end_of_day=False) if start_value else None
	)
	end_datetime = _parse_datetime_filter(end_value, end_of_day=True) if end_value else None

	if start_datetime and end_datetime and start_datetime > end_datetime:
		raise ValidationError({'detail': 'date_from cannot be after date_to.'})

	if start_datetime:
		queryset = queryset.filter(**{f'{field_name}__gte': start_datetime})
	if end_datetime:
		queryset = queryset.filter(**{f'{field_name}__lte': end_datetime})

	return queryset


def _apply_common_filters(
	queryset,
	request,
	*,
	date_field='created_at',
	status_field=None,
	warehouse_field=None,
	category_field=None,
):
	queryset = _apply_date_range_filter(queryset, request, field_name=date_field)

	status_value = _query_param(request, 'status')
	if status_field and status_value:
		queryset = queryset.filter(**{status_field: status_value})

	warehouse_value = _query_param(request, 'warehouse', 'warehouse_id')
	if warehouse_field and warehouse_value:
		queryset = queryset.filter(**{warehouse_field: warehouse_value})

	category_value = _query_param(request, 'category', 'category_id')
	if category_field and category_value:
		queryset = queryset.filter(**{category_field: category_value})

	return queryset


def _format_datetime_value(value, report_tz):
	if not value:
		return ''
	localized = _localize_datetime(value, report_tz)
	return localized.strftime('%Y-%m-%d %I:%M %p')


def _format_date_value(value):
	if not value:
		return ''
	localized = timezone.localtime(value)
	return localized.strftime('%Y-%m-%d')


def _resolve_report_timezone(request):
	timezone_name = _query_param(request, 'timezone', 'tz')
	if not timezone_name:
		return timezone.get_current_timezone()

	try:
		return ZoneInfo(timezone_name)
	except ZoneInfoNotFoundError as error:
		raise ValidationError({'detail': f'Unknown timezone: {timezone_name}'}) from error


def _localize_datetime(value, report_tz):
	if not isinstance(value, datetime):
		return value

	localized = value
	if timezone.is_naive(localized):
		localized = timezone.make_aware(localized, timezone.get_current_timezone())
	return localized.astimezone(report_tz)


def _export_timestamp():
	return timezone.localtime().strftime('%Y%m%d_%H%M%S')


def _normalize_status(value):
	if value in (None, ''):
		return ''
	return str(value).strip().lower().replace(' ', '_')


def _coerce_excel_cell_value(raw_value, value_type, report_tz):
	if raw_value is None:
		return '', None

	if value_type == 'date':
		if isinstance(raw_value, datetime):
			return _localize_datetime(raw_value, report_tz).date(), 'YYYY-MM-DD'
		if isinstance(raw_value, date):
			return raw_value, 'YYYY-MM-DD'
		if isinstance(raw_value, str):
			parsed_datetime = parse_datetime(raw_value)
			if parsed_datetime:
				return _localize_datetime(parsed_datetime, report_tz).date(), 'YYYY-MM-DD'
			parsed = parse_date(raw_value)
			if parsed:
				return parsed, 'YYYY-MM-DD'
		return raw_value, None

	if value_type == 'datetime':
		if isinstance(raw_value, datetime):
			localized = _localize_datetime(raw_value, report_tz)
			return localized.replace(tzinfo=None), 'YYYY-MM-DD HH:MM AM/PM'
		if isinstance(raw_value, str):
			parsed = parse_datetime(raw_value)
			if parsed:
				parsed = _localize_datetime(parsed, report_tz)
				return parsed.replace(tzinfo=None), 'YYYY-MM-DD HH:MM AM/PM'
		return raw_value, None

	if value_type == 'signed_number':
		try:
			return int(raw_value), '+0;-0;0'
		except (TypeError, ValueError):
			return raw_value, None

	if value_type == 'number':
		try:
			if isinstance(raw_value, bool):
				return int(raw_value), '0'
			if isinstance(raw_value, int):
				return raw_value, '0'
			if isinstance(raw_value, float):
				return raw_value, '0.00'
			string_value = str(raw_value).strip()
			if string_value == '':
				return '', None
			if '.' in string_value:
				return float(string_value), '0.00'
			return int(string_value), '0'
		except (TypeError, ValueError):
			return raw_value, None

	return raw_value, None


def _auto_adjust_column_width(worksheet, *, column_count, start_row, end_row):
	for column_index in range(1, column_count + 1):
		max_length = 0
		for row_index in range(start_row, end_row + 1):
			cell_value = worksheet.cell(row=row_index, column=column_index).value
			if cell_value in (None, ''):
				continue
			if isinstance(cell_value, datetime):
				rendered = cell_value.strftime('%Y-%m-%d %I:%M %p')
			elif isinstance(cell_value, date):
				rendered = cell_value.strftime('%Y-%m-%d')
			else:
				rendered = str(cell_value)
			max_length = max(max_length, len(rendered))

		column_letter = get_column_letter(column_index)
		worksheet.column_dimensions[column_letter].width = min(max(max_length + 5, 14), 60)


def generate_excel_report(data, columns, title, report_tz):
	workbook = Workbook()
	worksheet = workbook.active
	worksheet.title = title[:31]

	# Keep exported sheets printable as business PDFs without losing right-side columns.
	worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
	worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
	worksheet.page_setup.fitToWidth = 1
	worksheet.page_setup.fitToHeight = 0
	worksheet.page_margins = PageMargins(
		left=0.35,
		right=0.35,
		top=0.5,
		bottom=0.5,
		header=0.2,
		footer=0.2,
	)
	worksheet.print_options.horizontalCentered = True
	worksheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

	column_count = max(1, len(columns))
	worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
	title_cell = worksheet.cell(row=1, column=1, value=title)
	title_cell.font = TITLE_FONT
	title_cell.alignment = TITLE_ALIGNMENT
	worksheet.row_dimensions[1].height = 28

	generated_on = timezone.now().astimezone(report_tz).strftime('%Y-%m-%d %I:%M %p')
	generated_text = f'Generated on {generated_on}'
	worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=column_count)
	meta_cell = worksheet.cell(row=2, column=1, value=generated_text)
	meta_cell.font = META_FONT
	meta_cell.alignment = TITLE_ALIGNMENT

	header_row = 4
	worksheet.row_dimensions[header_row].height = 22
	for column_index, column in enumerate(columns, start=1):
		header_cell = worksheet.cell(row=header_row, column=column_index, value=column['label'])
		header_cell.font = HEADER_FONT
		header_cell.fill = HEADER_FILL
		header_cell.border = TABLE_BORDER
		header_cell.alignment = CENTER_ALIGNMENT

	data_start_row = header_row + 1
	for row_offset, row_data in enumerate(data):
		row_index = data_start_row + row_offset
		base_fill = ROW_FILL_EVEN if row_offset % 2 == 0 else ROW_FILL_ODD

		for column_index, column in enumerate(columns, start=1):
			value_type = column.get('type', 'text')
			raw_value = row_data.get(column['key'], '')
			cell_value, number_format = _coerce_excel_cell_value(raw_value, value_type, report_tz)

			cell = worksheet.cell(row=row_index, column=column_index, value=cell_value)
			cell.font = BODY_FONT
			cell.border = TABLE_BORDER
			cell.fill = base_fill
			cell.alignment = CENTER_ALIGNMENT

			if number_format:
				cell.number_format = number_format

			if value_type == 'status':
				status_key = _normalize_status(raw_value)
				status_style = STATUS_STYLE_MAP.get(status_key)
				if status_style:
					cell.fill = status_style['fill']
					cell.font = status_style['font']

	last_data_row = data_start_row + len(data) - 1 if data else header_row
	worksheet.freeze_panes = f'A{data_start_row}'
	worksheet.auto_filter.ref = (
		f'A{header_row}:{get_column_letter(column_count)}{last_data_row}'
	)
	worksheet.print_area = f'A1:{get_column_letter(column_count)}{max(last_data_row, header_row)}'
	worksheet.print_title_rows = f'{header_row}:{header_row}'

	_auto_adjust_column_width(
		worksheet,
		column_count=column_count,
		start_row=header_row,
		end_row=max(last_data_row, header_row),
	)

	buffer = BytesIO()
	workbook.save(buffer)
	buffer.seek(0)
	return buffer.getvalue()


def _build_excel_response(*, filename, title, columns, data, report_tz):
	content = generate_excel_report(data=data, columns=columns, title=title, report_tz=report_tz)

	response = HttpResponse(
		content,
		content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
	)
	response['Content-Disposition'] = f'attachment; filename="{filename}"'
	return response


def _build_ledger_pdf_response(*, filename, title, headers, rows, report_tz):
	buffer = BytesIO()
	document = SimpleDocTemplate(
		buffer,
		pagesize=landscape(A4),
		leftMargin=24,
		rightMargin=24,
		topMargin=24,
		bottomMargin=24,
	)
	styles = getSampleStyleSheet()
	title_style = ParagraphStyle(
		'LedgerTitle',
		parent=styles['Heading2'],
		fontName='Helvetica-Bold',
		fontSize=18,
		leading=22,
		textColor=colors.HexColor('#0f172a'),
		alignment=TA_CENTER,
	)
	meta_style = ParagraphStyle(
		'LedgerMeta',
		parent=styles['Normal'],
		fontName='Helvetica-Oblique',
		fontSize=10,
		leading=12,
		textColor=colors.HexColor('#475569'),
		alignment=TA_CENTER,
	)
	header_style = ParagraphStyle(
		'LedgerHeader',
		fontName='Helvetica-Bold',
		fontSize=10,
		leading=12,
		textColor=colors.whitesmoke,
		alignment=TA_CENTER,
	)
	body_left_style = ParagraphStyle(
		'LedgerBodyLeft',
		fontName='Helvetica',
		fontSize=9.5,
		leading=12,
		textColor=colors.HexColor('#111827'),
		alignment=TA_LEFT,
	)
	body_center_style = ParagraphStyle(
		'LedgerBodyCenter',
		fontName='Helvetica',
		fontSize=9.5,
		leading=12,
		textColor=colors.HexColor('#111827'),
		alignment=TA_CENTER,
	)

	table_rows = [[Paragraph(escape(str(item)), header_style) for item in headers]]
	for row in rows:
		table_rows.append(
			[
				Paragraph(escape(str(row[0] or '')), body_left_style),
				Paragraph(escape(str(row[1] or '')), body_center_style),
				Paragraph(escape(str(row[2] or '')), body_center_style),
				Paragraph(escape(str(row[3] or '')), body_left_style),
				Paragraph(escape(str(row[4] or '')), body_left_style),
				Paragraph(escape(str(row[5] or '')), body_center_style),
			]
		)

	table = Table(
		table_rows,
		repeatRows=1,
		colWidths=[122, 92, 110, 183, 170, 116],
		hAlign='CENTER',
	)
	table.setStyle(
		TableStyle(
			[
				('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
				('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
				('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
				('FONTSIZE', (0, 0), (-1, -1), 9.5),
				('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#d1d5db')),
				('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
				('LEFTPADDING', (0, 0), (-1, -1), 6),
				('RIGHTPADDING', (0, 0), (-1, -1), 6),
				('TOPPADDING', (0, 0), (-1, -1), 4),
				('BOTTOMPADDING', (0, 0), (-1, -1), 4),
				('ALIGN', (1, 1), (2, -1), 'CENTER'),
				('ALIGN', (5, 1), (5, -1), 'CENTER'),
				('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
			]
		)
	)

	generated_at = timezone.now().astimezone(report_tz).strftime('%Y-%m-%d %I:%M %p')

	document.build(
		[
			Paragraph(escape(title), title_style),
			Paragraph(f'Generated on {generated_at}', meta_style),
			Spacer(1, 8),
			table,
		]
	)

	response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
	response['Content-Disposition'] = f'attachment; filename="{filename}"'
	return response


def _apply_staff_export_limit(queryset, user):
	if user.is_superuser or user.role != User.Roles.STAFF:
		return queryset
	return queryset.filter(created_by=user)


def _location_label(location):
	if location is None:
		return '-'
	warehouse = getattr(location, 'warehouse', None)
	if warehouse is not None:
		return f'{warehouse.code} - {location.name}'
	return location.name


class BaseExportView(APIView):
	permission_classes = [IsAuthenticated]

	def no_data_response(self):
		return Response({'detail': NO_DATA_DETAIL}, status=status.HTTP_400_BAD_REQUEST)


class ProductExportView(BaseExportView):
	permission_classes = [IsAuthenticated, IsManagerOrSuperUser]

	def get(self, request):
		report_tz = _resolve_report_timezone(request)
		warehouse_value = _query_param(request, 'warehouse', 'warehouse_id')

		queryset = Product.objects.select_related('category').all()
		queryset = _apply_common_filters(
			queryset,
			request,
			date_field='created_at',
			warehouse_field='stock_balances__location__warehouse_id',
			category_field='category_id',
		)

		search = _query_param(request, 'search', 'q')
		if search:
			queryset = queryset.filter(
				Q(name__icontains=search)
				| Q(sku__icontains=search)
				| Q(category__name__icontains=search)
			)

		if warehouse_value:
			queryset = queryset.annotate(
				warehouse_stock=Coalesce(
					Sum(
						'stock_balances__quantity',
						filter=Q(stock_balances__location__warehouse_id=warehouse_value),
					),
					0,
				)
			)

		queryset = queryset.order_by('name').distinct()

		records = []
		for product in queryset:
			stock_value = product.current_stock
			if warehouse_value:
				stock_value = getattr(product, 'warehouse_stock', 0)
			records.append(
				{
					'product_name': product.name,
					'sku': product.sku,
					'category': product.category.name if product.category else '-',
					'unit': product.unit_of_measure,
					'stock': int(stock_value or 0),
					'reorder_level': int(product.reorder_level or 0),
				}
			)

		if not records:
			return self.no_data_response()

		filename = f'products_export_{_export_timestamp()}.xlsx'
		return _build_excel_response(
			filename=filename,
			title='Products Report',
			columns=[
				{'key': 'product_name', 'label': 'Product Name', 'type': 'text'},
				{'key': 'sku', 'label': 'SKU', 'type': 'text'},
				{'key': 'category', 'label': 'Category', 'type': 'text'},
				{'key': 'unit', 'label': 'Unit', 'type': 'text'},
				{'key': 'stock', 'label': 'Stock', 'type': 'number'},
				{'key': 'reorder_level', 'label': 'Reorder Level', 'type': 'number'},
			],
			data=records,
			report_tz=report_tz,
		)


class LedgerExportView(BaseExportView):
	permission_classes = [IsAuthenticated, IsManagerOrSuperUser]

	def get(self, request):
		report_tz = _resolve_report_timezone(request)
		export_format = (
			_query_param(request, 'file_format', 'export_format', 'format') or 'xlsx'
		).lower()
		if export_format not in {'xlsx', 'pdf'}:
			raise ValidationError({'detail': 'format must be either xlsx or pdf.'})

		queryset = StockLedger.objects.select_related(
			'product',
			'location',
			'location__warehouse',
		).all()
		queryset = _apply_date_range_filter(queryset, request, field_name='created_at')

		movement_type = _query_param(request, 'movement_type', 'status')
		if movement_type:
			queryset = queryset.filter(movement_type=movement_type)

		warehouse_value = _query_param(request, 'warehouse', 'warehouse_id')
		if warehouse_value:
			queryset = queryset.filter(location__warehouse_id=warehouse_value)

		category_value = _query_param(request, 'category', 'category_id')
		if category_value:
			queryset = queryset.filter(product__category_id=category_value)

		product_value = _query_param(request, 'product')
		if product_value:
			queryset = queryset.filter(product_id=product_value)

		location_value = _query_param(request, 'location')
		if location_value:
			queryset = queryset.filter(location_id=location_value)

		search = _query_param(request, 'search', 'q')
		if search:
			queryset = queryset.filter(
				Q(reference_id__icontains=search)
				| Q(note__icontains=search)
				| Q(product__name__icontains=search)
				| Q(product__sku__icontains=search)
			)

		queryset = queryset.order_by('-created_at')

		records = []
		pdf_rows = []
		for entry in queryset:
			operation_label = entry.get_movement_type_display()
			location_label = _location_label(entry.location)
			records.append(
				{
					'product': entry.product.name,
					'quantity_change': entry.change,
					'operation_type': operation_label,
					'reference_id': entry.reference_id,
					'location': location_label,
					'date_time': entry.created_at,
				}
			)
			pdf_rows.append(
				(
					entry.product.name,
					f'{entry.change:+d}',
					operation_label,
					entry.reference_id,
					location_label,
					_format_datetime_value(entry.created_at, report_tz),
				)
			)

		if not records:
			return self.no_data_response()

		if export_format == 'pdf':
			filename = f'stock_ledger_{_export_timestamp()}.pdf'
			return _build_ledger_pdf_response(
				filename=filename,
				title='Stock Ledger Report',
				headers=[
					'Product',
					'Quantity Change',
					'Operation Type',
					'Reference ID',
					'Location',
					'Date & Time',
				],
				rows=pdf_rows,
				report_tz=report_tz,
			)

		filename = f'stock_ledger_{_export_timestamp()}.xlsx'
		return _build_excel_response(
			filename=filename,
			title='Stock Ledger Report',
			columns=[
				{'key': 'product', 'label': 'Product', 'type': 'text'},
				{'key': 'quantity_change', 'label': 'Quantity Change', 'type': 'signed_number'},
				{'key': 'operation_type', 'label': 'Operation Type', 'type': 'text'},
				{'key': 'reference_id', 'label': 'Reference ID', 'type': 'text'},
				{'key': 'location', 'label': 'Location', 'type': 'text'},
				{'key': 'date_time', 'label': 'Date & Time', 'type': 'datetime'},
			],
			data=records,
			report_tz=report_tz,
		)


class ReceiptsExportView(BaseExportView):
	permission_classes = [IsAuthenticated, IsManagerOrStaff]

	def get(self, request):
		report_tz = _resolve_report_timezone(request)
		queryset = Receipt.objects.select_related(
			'supplier',
			'destination_location',
			'destination_location__warehouse',
			'created_by',
		).prefetch_related('items__product')
		queryset = _apply_common_filters(
			queryset,
			request,
			date_field='created_at',
			status_field='status',
			warehouse_field='destination_location__warehouse_id',
			category_field='items__product__category_id',
		)

		search = _query_param(request, 'search', 'q')
		if search:
			queryset = queryset.filter(
				Q(reference_no__icontains=search)
				| Q(supplier__name__icontains=search)
				| Q(items__product__name__icontains=search)
				| Q(items__product__sku__icontains=search)
			)

		queryset = _apply_staff_export_limit(queryset, request.user)
		queryset = queryset.order_by('-created_at').distinct()

		records = []
		for receipt in queryset:
			items = list(receipt.items.all())
			if not items:
				records.append(
					{
						'receipt_id': receipt.reference_no,
						'supplier': receipt.supplier.name,
						'product': '-',
						'quantity': 0,
						'status': receipt.get_status_display(),
						'date': receipt.created_at,
					}
				)
				continue

			for item in items:
				records.append(
					{
						'receipt_id': receipt.reference_no,
						'supplier': receipt.supplier.name,
						'product': item.product.name,
						'quantity': item.quantity,
						'status': receipt.get_status_display(),
						'date': receipt.created_at,
					}
				)

		if not records:
			return self.no_data_response()

		filename = f'receipts_export_{_export_timestamp()}.xlsx'
		return _build_excel_response(
			filename=filename,
			title='Receipts Report',
			columns=[
				{'key': 'receipt_id', 'label': 'Receipt ID', 'type': 'text'},
				{'key': 'supplier', 'label': 'Supplier', 'type': 'text'},
				{'key': 'product', 'label': 'Product', 'type': 'text'},
				{'key': 'quantity', 'label': 'Quantity', 'type': 'number'},
				{'key': 'status', 'label': 'Status', 'type': 'status'},
				{'key': 'date', 'label': 'Date', 'type': 'date'},
			],
			data=records,
			report_tz=report_tz,
		)


class DeliveriesExportView(BaseExportView):
	permission_classes = [IsAuthenticated, IsManagerOrStaff]

	def get(self, request):
		report_tz = _resolve_report_timezone(request)
		queryset = Delivery.objects.select_related(
			'customer',
			'source_location',
			'source_location__warehouse',
			'created_by',
		).prefetch_related('items__product')
		queryset = _apply_common_filters(
			queryset,
			request,
			date_field='created_at',
			status_field='status',
			warehouse_field='source_location__warehouse_id',
			category_field='items__product__category_id',
		)

		search = _query_param(request, 'search', 'q')
		if search:
			queryset = queryset.filter(
				Q(reference_no__icontains=search)
				| Q(customer__name__icontains=search)
				| Q(items__product__name__icontains=search)
				| Q(items__product__sku__icontains=search)
			)

		queryset = _apply_staff_export_limit(queryset, request.user)
		queryset = queryset.order_by('-created_at').distinct()

		records = []
		for delivery in queryset:
			items = list(delivery.items.all())
			if not items:
				records.append(
					{
						'delivery_id': delivery.reference_no,
						'customer': delivery.customer.name,
						'product': '-',
						'quantity': 0,
						'status': delivery.get_status_display(),
						'date': delivery.created_at,
					}
				)
				continue

			for item in items:
				records.append(
					{
						'delivery_id': delivery.reference_no,
						'customer': delivery.customer.name,
						'product': item.product.name,
						'quantity': item.quantity,
						'status': delivery.get_status_display(),
						'date': delivery.created_at,
					}
				)

		if not records:
			return self.no_data_response()

		filename = f'deliveries_export_{_export_timestamp()}.xlsx'
		return _build_excel_response(
			filename=filename,
			title='Deliveries Report',
			columns=[
				{'key': 'delivery_id', 'label': 'Delivery ID', 'type': 'text'},
				{'key': 'customer', 'label': 'Customer', 'type': 'text'},
				{'key': 'product', 'label': 'Product', 'type': 'text'},
				{'key': 'quantity', 'label': 'Quantity', 'type': 'number'},
				{'key': 'status', 'label': 'Status', 'type': 'status'},
				{'key': 'date', 'label': 'Date', 'type': 'date'},
			],
			data=records,
			report_tz=report_tz,
		)


class TransfersExportView(BaseExportView):
	permission_classes = [IsAuthenticated, IsManagerOrStaff]

	def get(self, request):
		report_tz = _resolve_report_timezone(request)
		queryset = InternalTransfer.objects.select_related(
			'from_location',
			'to_location',
			'from_location__warehouse',
			'to_location__warehouse',
			'created_by',
		).prefetch_related('items__product')
		queryset = _apply_common_filters(
			queryset,
			request,
			date_field='created_at',
			status_field='status',
			category_field='items__product__category_id',
		)

		warehouse_value = _query_param(request, 'warehouse', 'warehouse_id')
		if warehouse_value:
			queryset = queryset.filter(
				Q(from_location__warehouse_id=warehouse_value)
				| Q(to_location__warehouse_id=warehouse_value)
			)

		search = _query_param(request, 'search', 'q')
		if search:
			queryset = queryset.filter(
				Q(reference_no__icontains=search)
				| Q(items__product__name__icontains=search)
				| Q(items__product__sku__icontains=search)
			)

		queryset = _apply_staff_export_limit(queryset, request.user)
		queryset = queryset.order_by('-created_at').distinct()

		records = []
		for transfer in queryset:
			items = list(transfer.items.all())
			if not items:
				records.append(
					{
						'transfer_id': transfer.reference_no,
						'from_location': _location_label(transfer.from_location),
						'to_location': _location_label(transfer.to_location),
						'product': '-',
						'quantity': 0,
						'status': transfer.get_status_display(),
						'date': transfer.created_at,
					}
				)
				continue

			for item in items:
				records.append(
					{
						'transfer_id': transfer.reference_no,
						'from_location': _location_label(transfer.from_location),
						'to_location': _location_label(transfer.to_location),
						'product': item.product.name,
						'quantity': item.quantity,
						'status': transfer.get_status_display(),
						'date': transfer.created_at,
					}
				)

		if not records:
			return self.no_data_response()

		filename = f'transfers_export_{_export_timestamp()}.xlsx'
		return _build_excel_response(
			filename=filename,
			title='Transfers Report',
			columns=[
				{'key': 'transfer_id', 'label': 'Transfer ID', 'type': 'text'},
				{'key': 'from_location', 'label': 'From Location', 'type': 'text'},
				{'key': 'to_location', 'label': 'To Location', 'type': 'text'},
				{'key': 'product', 'label': 'Product', 'type': 'text'},
				{'key': 'quantity', 'label': 'Quantity', 'type': 'number'},
				{'key': 'status', 'label': 'Status', 'type': 'status'},
				{'key': 'date', 'label': 'Date', 'type': 'date'},
			],
			data=records,
			report_tz=report_tz,
		)


class AdjustmentsExportView(BaseExportView):
	permission_classes = [IsAuthenticated, IsManagerOrSuperUser]

	def get(self, request):
		report_tz = _resolve_report_timezone(request)
		queryset = StockAdjustment.objects.select_related(
			'product',
			'location',
			'location__warehouse',
		).all()
		queryset = _apply_common_filters(
			queryset,
			request,
			date_field='created_at',
			status_field='status',
			warehouse_field='location__warehouse_id',
			category_field='product__category_id',
		)

		search = _query_param(request, 'search', 'q')
		if search:
			queryset = queryset.filter(
				Q(reference_no__icontains=search)
				| Q(product__name__icontains=search)
				| Q(product__sku__icontains=search)
				| Q(notes__icontains=search)
			)

		queryset = queryset.order_by('-created_at')

		records = [
			{
				'adjustment_id': adjustment.reference_no,
				'product': adjustment.product.name,
				'system_qty': adjustment.system_quantity,
				'counted_qty': adjustment.counted_quantity,
				'difference': adjustment.difference,
				'reason': adjustment.notes or '-',
				'date': adjustment.created_at,
			}
			for adjustment in queryset
		]

		if not records:
			return self.no_data_response()

		filename = f'adjustments_export_{_export_timestamp()}.xlsx'
		return _build_excel_response(
			filename=filename,
			title='Adjustments Report',
			columns=[
				{'key': 'adjustment_id', 'label': 'Adjustment ID', 'type': 'text'},
				{'key': 'product', 'label': 'Product', 'type': 'text'},
				{'key': 'system_qty', 'label': 'System Qty', 'type': 'number'},
				{'key': 'counted_qty', 'label': 'Counted Qty', 'type': 'number'},
				{'key': 'difference', 'label': 'Difference', 'type': 'signed_number'},
				{'key': 'reason', 'label': 'Reason', 'type': 'text'},
				{'key': 'date', 'label': 'Date', 'type': 'date'},
			],
			data=records,
			report_tz=report_tz,
		)